

import os
import io
import json
import re
import time
import csv
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List
from urllib.parse import quote

import streamlit as st
import pdfplumber
from docx import Document
from dotenv import load_dotenv
from openpyxl import Workbook

from google import genai


# ----------------------------
# Configuration
# ----------------------------
CSV_PATH = "outreach_log.csv"
XLSX_PATH = "outreach_log.xlsx"
SHEET_NAME = "Outreach"

# CSV columns you wanted
HEADERS = ["Person Name", "Company", "Role", "Person Email", "Date", "Resume File"]

RESUME_SAVE_DIR = "saved_resumes"

# Keep prompts smaller to reduce cost/quota
MAX_PROFILE_CHARS = 1500
MAX_JD_CHARS = 3500
MAX_RESUME_CHARS = 6500

DEFAULT_GEMINI_MODEL = os.getenv("GEMINI_MODEL", "gemini-2.0-flash")

CANDIDATE_NAME_DEFAULT = "Shivakumar Pasem"


# ----------------------------
# Utility
# ----------------------------
def clamp(s: str, n: int) -> str:
    return (s or "")[:n]

def safe_slug(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"[^a-zA-Z0-9_.-]+", "_", s)
    return s[:120] if s else "file"

def recipient_name_from_header(profile_header: str) -> str:
    """
    Extract a clean recipient name from messy LinkedIn copy-paste.
    Handles cases like:
    "Aubre Camp 3rd degree connection3rd\"Strive not...\" New Age Industrial ..."
    """
    s = (profile_header or "").strip()
    if not s:
        return ""

    # 1) Use only the first line
    s = s.splitlines()[0].strip()

    # 2) If there's a pipe format, take left side
    if "|" in s:
        s = s.split("|")[0].strip()

    # 3) Remove common LinkedIn noise
    # Remove connection indicators like "3rd degree connection3rd", "1st degree connection", etc.
    s = re.sub(r"\b(1st|2nd|3rd)\s*degree\s*connection\d*\b", "", s, flags=re.IGNORECASE)
    s = re.sub(r"\bconnection\d*\b", "", s, flags=re.IGNORECASE)

    # Remove quoted tagline chunks (often long quotes)
    s = re.sub(r"\".*?\"", "", s)

    # Remove separators
    s = s.replace("•", " ").replace("·", " ").replace("—", " ").replace("-", " ")
    s = re.sub(r"\s+", " ", s).strip()

    # 4) Now grab name-like tokens from the beginning
    tokens = s.split()
    cleaned: List[str] = []
    for t in tokens:
        # Stop when we hit a token that signals the start of non-name info
        if t.lower() in {"new", "age", "industrial", "university", "college", "inc", "llc"}:
            break
        if "," in t:
            break

        # Keep only alphabetic-ish tokens (names)
        t2 = re.sub(r"[^A-Za-z.'-]", "", t)
        if not t2:
            continue

        cleaned.append(t2)

        # Most names are 2-3 words, allow 4 max
        if len(cleaned) >= 4:
            break

    # Need at least a first name
    if len(cleaned) == 0:
        return ""

    # If only 1 token and it looks suspiciously short, return ""
    if len(cleaned) == 1 and len(cleaned[0]) < 3:
        return ""

    return " ".join(cleaned)


def recipient_first_name(profile_header: str) -> str:
    full = recipient_name_from_header(profile_header)
    if not full:
        return "there"
    parts = full.split()
    return parts[0] if parts else "there"

def file_path_to_uri(abs_path: str) -> str:
    p = abs_path.replace("\\", "/")
    return "file:///" + quote(p)

def ensure_csv(path: str) -> None:
    if os.path.exists(path):
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(HEADERS)

def append_csv_row(path: str, row: List[Any]) -> None:
    ensure_csv(path)
    with open(path, "a", newline="", encoding="utf-8") as f:
        w = csv.writer(f)
        w.writerow(row)

def read_csv_rows(path: str) -> List[List[str]]:
    if not os.path.exists(path):
        return []
    rows: List[List[str]] = []
    with open(path, "r", newline="", encoding="utf-8") as f:
        r = csv.reader(f)
        for i, row in enumerate(r):
            if i == 0:
                continue
            if row:
                rows.append(row)
    return rows

def export_csv_to_excel(csv_path: str, xlsx_path: str) -> None:
    ensure_csv(csv_path)
    rows = read_csv_rows(csv_path)

    wb = Workbook()
    ws = wb.active
    ws.title = SHEET_NAME
    ws.append(HEADERS)

    for row in rows:
        while len(row) < len(HEADERS):
            row.append("")
        person_name, company, role, email, date_str, resume_ref = row[:6]

        ws.append([person_name, company, role, email, date_str, ""])
        resume_cell = ws.cell(row=ws.max_row, column=6)

        resume_ref = (resume_ref or "").strip()
        if resume_ref.lower().startswith("file:///"):
            display = resume_ref.split("/")[-1]
            resume_cell.value = display
            resume_cell.hyperlink = resume_ref
            resume_cell.style = "Hyperlink"
        else:
            resume_cell.value = resume_ref

    wb.save(xlsx_path)


# ----------------------------
# Resume handling
# Save only when you log to CSV
# ----------------------------
def save_resume_bytes_on_log(resume_bytes: bytes, original_filename: str, version_tag: str) -> str:
    """
    Save the uploaded resume only at logging time, then return absolute path.
    """
    os.makedirs(RESUME_SAVE_DIR, exist_ok=True)

    stamp = datetime.now().strftime("%Y-%m-%d")
    version = safe_slug(version_tag or "v1")
    original = safe_slug(original_filename or "resume.pdf")

    filename = f"{stamp}__{version}__{original}"
    abs_path = os.path.abspath(os.path.join(RESUME_SAVE_DIR, filename))

    with open(abs_path, "wb") as f:
        f.write(resume_bytes)

    return abs_path

def extract_text_from_pdf_bytes(pdf_bytes: bytes) -> str:
    parts: List[str] = []
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        for page in pdf.pages:
            t = page.extract_text() or ""
            if t.strip():
                parts.append(t)
    return "\n".join(parts).strip()

def extract_text_from_docx_bytes(docx_bytes: bytes) -> str:
    doc = Document(io.BytesIO(docx_bytes))
    parts = [p.text for p in doc.paragraphs if p.text.strip()]
    return "\n".join(parts).strip()


# ----------------------------
# Gemini + JSON handling
# ----------------------------
def extract_json_block(text: str) -> Dict[str, Any]:
    text = (text or "").strip()
    try:
        return json.loads(text)
    except Exception:
        pass
    m = re.search(r"\{.*\}", text, re.DOTALL)
    if not m:
        raise ValueError("Model output did not contain a JSON object.")
    return json.loads(m.group(0))

def gemini_generate(prompt: str, model: str) -> str:
    client = genai.Client()
    resp = client.models.generate_content(model=model, contents=prompt)
    return resp.text or ""

def gemini_generate_with_retry(prompt: str, model: str, max_tries: int = 2) -> str:
    last_err = None
    for _ in range(max_tries):
        try:
            return gemini_generate(prompt, model)
        except Exception as e:
            last_err = e
            msg = str(e)
            if "429" in msg or "RESOURCE_EXHAUSTED" in msg:
                time.sleep(10)
                continue
            break
    raise last_err  # type: ignore


# ----------------------------
# Prompt builder
# Fixes:
# - Recipient name never comes from the model.
# - Email is written from candidate perspective ("I"), not recruiter voice.
# - We generate paragraphs, then assemble exact template in Python.
# - Also generates an ATS rewrite prompt you can paste into ChatGPT later.
# ----------------------------
def build_prompt(profile_header: str, jd: str, resume_text: str) -> str:
    return f"""
Return ONLY valid JSON with exactly these keys. No extra text.

Keys:
company: string or null
role: string or null

score_out_of_10: number (0 to 10)
score_reason: string (2 to 4 lines max)
top_matches: array of exactly 3 strings (evidence from resume)
gaps: array of exactly 3 strings (missing or weak vs JD)
resume_tweaks: array of exactly 4 strings (specific, truthful bullet rewrite ideas)

cold_email_subject: string
email_p1_intro_interest: string (2 to 3 sentences)
email_p2_experience_alignment: string (2 to 3 sentences)
email_p3_cta_resume: string (1 to 2 sentences)

linkedin_note_under_300: string (under 300 characters)

Hard rules:
- Act as a hiring manager scoring the candidate for THIS job.
- Use ONLY the resume text as facts about the candidate. Do not invent tools, employers, metrics.
- Email must be written from the candidate perspective using "I" and "my".
- Do NOT write from recruiter perspective. Avoid phrases like "I noticed your profile" or "your background caught my eye".
- Do NOT greet or address the candidate. Recipient is the LinkedIn person.
- No bullets in email paragraphs.
- No em dashes.

Email content rules:
- Subject: catchy but professional, include the role plus one hook from JD or profile header.
- Paragraph 1: include candidate name "Shivakumar Pasem", state interest in the role and company, and one specific reason based on JD or profile header.
- Paragraph 2: mention 4 to 6 relevant skills from resume only and connect to JD.
- Paragraph 3: ask for a brief chat at their convenience and mention resume attached.

LinkedIn note rules:
- Must sound like outreach from candidate, not recruiter.
- Include one specific hook from JD or profile header and one proof point from resume.
- End with a soft CTA.
- Under 300 characters.

Inputs:
LinkedIn header (recipient context):
{profile_header}

Job description:
{jd}

Resume text:
{resume_text}
""".strip()


# ----------------------------
# Template assembly (guarantees exact format)
# ----------------------------
def assemble_email(recipient_first: str, p1: str, p2: str, p3: str, signature: str) -> str:
    recipient_first = (recipient_first or "there").strip() or "there"
    signature = (signature or CANDIDATE_NAME_DEFAULT).strip() or CANDIDATE_NAME_DEFAULT

    p1 = " ".join((p1 or "").splitlines()).strip()
    p2 = " ".join((p2 or "").splitlines()).strip()
    p3 = " ".join((p3 or "").splitlines()).strip()

    return (
        f"Hi {recipient_first},\n\n"
        f"{p1}\n\n"
        f"{p2}\n\n"
        f"{p3}\n\n"
        f"Best regards,\n"
        f"{signature}"
    )

def normalize_linkedin_note(note: str, recipient_first: str) -> str:
    note = (note or "").strip()
    recipient_first = (recipient_first or "").strip()

    if recipient_first and recipient_first.lower() != "there":
        if not note.lower().startswith(("hi ", "hello ")):
            note = f"Hi {recipient_first}, {note}"
    else:
        if not note.lower().startswith(("hi ", "hello ")):
            note = f"Hi, {note}"

    if len(note) > 300:
        note = note[:297].rstrip() + "..."
    return note

def build_ats_rewrite_prompt(
    jd: str,
    role: str,
    company: str,
    top_matches: List[str],
    gaps: List[str],
    tweaks: List[str]
) -> str:
    """
    Creates a prompt you can paste into ChatGPT along with your resume to do ATS-aligned rewrites.
    Note: your Google Doc could not be programmatically opened (Google Docs blocks non-JS viewers),
    so this follows strong ATS norms: truthful, keyword-aligned, clean formatting, impact-first bullets.
    """
    role = (role or "the target role").strip()
    company = (company or "the company").strip()

    def bullets(items: List[str], max_n: int) -> str:
        items = [x.strip() for x in (items or []) if x and x.strip()]
        items = items[:max_n]
        return "\n".join([f"- {x}" for x in items]) if items else "- (none provided)"

    return f"""
You are an expert resume writer and ATS optimization specialist.

I will paste my current resume after this prompt.

Goal:
Rewrite my resume to be ATS-optimized and tailored for {role} at {company}. Keep it truthful and consistent with my original resume. Do not invent employers, tools, certifications, degrees, or metrics.

Strict rules:

2) Keyword alignment:
- Extract the most important hard skills, tools, and role keywords from the JD.
- Integrate keywords naturally across Summary, Skills, and bullets.
- Avoid keyword stuffing.

3) Bullet quality:
- Prefer impact-first bullets: action + what + outcome.
- Use strong verbs (improved, automated, delivered, optimized, resolved).
- Replace vague phrases like "responsible for" with specific actions.
- If a metric is not in my resume, do not make one up. You may rephrase to show impact without numbers.

4) Relevance ordering:
- Put the most relevant experience and bullets first within each role.
- Keep only role-relevant content. Remove weak filler.

Use this analysis guidance:
Top matches to preserve:
{bullets(top_matches, 3)}

Gaps to address:
{bullets(gaps, 3)}

High-priority tweaks to implement:
{bullets(tweaks, 6)}

Job Description:
\"\"\"{jd.strip()}\"\"\"

Now ask me to paste my resume. After I paste it, output the complete revised resume as clean text.
""".strip()


# ----------------------------
# Streamlit UI
# ----------------------------
st.set_page_config(page_title="Resume Agent", layout="wide")
st.title("Resume Agent")

load_dotenv(dotenv_path=Path(__file__).with_name(".env"))
if not os.getenv("GEMINI_API_KEY"):
    st.error("GEMINI_API_KEY not found. Add it to .env next to app.py and restart Streamlit.")
    st.stop()

colL, colR = st.columns([1, 1.2])

with colL:
    st.subheader("Inputs")
    model = st.text_input("Gemini model", value=DEFAULT_GEMINI_MODEL)

    profile = st.text_area(
        "Paste LinkedIn profile header (recipient)",
        height=120,
        placeholder="Example: Britt Smith | Recruiting at Indeavor"
    )
    jd = st.text_area("Paste Job Description", height=220)

    resume_file = st.file_uploader("Upload Resume (PDF or DOCX)", type=["pdf", "docx"])
    resume_version = st.text_input("Resume version tag (example: v1, v2, DA_v3)", value="v1")
    signature_name = st.text_input("Signature name", value=CANDIDATE_NAME_DEFAULT)

    # Recipient name derived only from LinkedIn header
    recipient_full = recipient_name_from_header(profile)
    recipient_first = recipient_first_name(profile)

    if recipient_first.lower() in {"shivakumar", "pasem"}:
        st.warning("Recipient name looks like your name. Double-check you pasted the recipient's LinkedIn header.")
        recipient_first = "there"
    st.caption(f"Will log person name as: {recipient_full or '(not detected)'}")
    # st.caption(f"Recipient parsed as: {recipient_full or '(not found)'}")

    # Keep resume bytes in memory for later logging save
    resume_bytes: bytes = b""
    resume_filename: str = ""
    resume_text = ""

    if resume_file:
        resume_filename = resume_file.name
        resume_bytes = resume_file.getvalue()

        try:
            if resume_filename.lower().endswith(".pdf"):
                resume_text = extract_text_from_pdf_bytes(resume_bytes)
            else:
                resume_text = extract_text_from_docx_bytes(resume_bytes)
        except Exception as e:
            st.error(f"Could not read resume file: {e}")

    st.subheader("Resume Preview (first 800 chars)")
    st.write(resume_text[:800] if resume_text else "Upload a resume to preview extracted text.")

    # Trim inputs for quota and speed
    profile_trim = clamp(profile, MAX_PROFILE_CHARS)
    jd_trim = clamp(jd, MAX_JD_CHARS)
    resume_trim = clamp(resume_text, MAX_RESUME_CHARS)

    st.caption(f"Trimmed sizes: profile {len(profile_trim)}, JD {len(jd_trim)}, resume {len(resume_trim)} chars.")

    can_run = bool(profile_trim.strip()) and bool(jd_trim.strip()) and bool(resume_trim.strip())

    generate_clicked = st.button(
        "Generate Score, Tweaks, Cold Email, LinkedIn Note",
        type="primary",
        disabled=not can_run
    )

    st.divider()
    st.subheader("Logging and Export")

    if st.button("Export CSV to Excel (.xlsx)"):
        export_csv_to_excel(CSV_PATH, XLSX_PATH)
        st.success(f"Exported to {XLSX_PATH}. If Excel is open, close it before exporting again.")

with colR:
    st.subheader("Outputs")

    if generate_clicked:
        prompt = build_prompt(profile_trim, jd_trim, resume_trim)

        with st.spinner("Generating..."):
            raw = gemini_generate_with_retry(prompt, model=model, max_tries=2)
            data = extract_json_block(raw)

        # Persist everything needed for logging later
        st.session_state["data"] = data
        st.session_state["recipient_full"] = recipient_full
        st.session_state["recipient_first"] = recipient_first
        st.session_state["resume_bytes"] = resume_bytes
        st.session_state["resume_filename"] = resume_filename
        st.session_state["resume_version"] = resume_version
        st.session_state["signature_name"] = signature_name
        st.session_state["jd_trim"] = jd_trim

    data = st.session_state.get("data")
    if data:
        recipient_full = st.session_state.get("recipient_full", "")
        recipient_first = st.session_state.get("recipient_first", "there")
        signature_name = st.session_state.get("signature_name", CANDIDATE_NAME_DEFAULT)
        jd_trim = st.session_state.get("jd_trim", "")

        company = (data.get("company") or "").strip()
        role = (data.get("role") or "").strip()

        score = data.get("score_out_of_10", "")
        score_reason = data.get("score_reason", "")

        top_matches = data.get("top_matches") or []
        gaps = data.get("gaps") or []
        tweaks = data.get("resume_tweaks") or []

        subject = (data.get("cold_email_subject") or "").strip()
        p1 = data.get("email_p1_intro_interest") or ""
        p2 = data.get("email_p2_experience_alignment") or ""
        p3 = data.get("email_p3_cta_resume") or ""

        note_raw = (data.get("linkedin_note_under_300") or "").strip()

        email_body = assemble_email(recipient_first, p1, p2, p3, signature_name)
        linkedin_note = normalize_linkedin_note(note_raw, recipient_first)

        ats_prompt = build_ats_rewrite_prompt(
            jd=jd_trim,
            role=role,
            company=company,
            top_matches=top_matches,
            gaps=gaps,
            tweaks=tweaks
        )

        st.markdown("### Recipient")
        st.write(recipient_full or "(not found)")

        st.markdown("### Hiring Manager Fit Score")
        st.write(f"Score: {score} / 10")
        st.text(score_reason)

        st.markdown("### Top Matches")
        for x in top_matches[:3]:
            st.write(f"- {x}")

        st.markdown("### Gaps")
        for x in gaps[:3]:
            st.write(f"- {x}")

        st.markdown("### Resume Tweaks")
        for x in tweaks[:4]:
            st.write(f"- {x}")

        st.markdown("### Cold Email")
        if subject:
            st.write(f"Subject: {subject}")
        st.text(email_body)

        st.markdown("### LinkedIn Connection Note")
        st.text(linkedin_note)
        st.caption(f"Characters: {len(linkedin_note)}")

        st.markdown("### ATS Resume Rewrite Prompt (paste into ChatGPT)")
        st.text_area("Copy this prompt", value=ats_prompt, height=320)

        st.divider()
        st.markdown("### Done and Log to CSV")

        person_email_input = st.text_input("Person Email (optional)", value="")

        if st.button("Done, Log to CSV"):
            # Save resume only now (because it is being logged)
            resume_bytes = st.session_state.get("resume_bytes", b"")
            resume_filename = st.session_state.get("resume_filename", "")
            resume_version = st.session_state.get("resume_version", "v1")

            if not resume_bytes or not resume_filename:
                st.error("Please upload a resume before logging.")
                st.stop()

            try:
                saved_abs = save_resume_bytes_on_log(resume_bytes, resume_filename, resume_version)
                saved_uri = file_path_to_uri(saved_abs)
                resume_ref = saved_uri
            except Exception as e:
                st.warning(f"Could not save resume locally. Logging version tag instead. Error: {e}")
                resume_ref = resume_version

            append_csv_row(
                CSV_PATH,
                [
                    recipient_full,                 # Person Name (recipient)
                    company,                         # Company
                    role,                            # Role
                    person_email_input or "",        # Person Email
                    datetime.now().strftime("%Y-%m-%d"),
                    resume_ref                       # Resume File (file:///... or version)
                ],
            )
            st.success(f"Logged to {CSV_PATH}. Resume saved only for logged entries.")
    else:
        st.info("Generate outputs to see score, email, note, ATS prompt, and logging controls.")
